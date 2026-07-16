#!/usr/bin/env python3
"""Human Evaluation Sampling Script for Experiment 1.

Samples ~100 cases (stratified by LLM Judge OA score) from DeepSeek-V4-Flash
generated responses and produces an Excel spreadsheet for human annotators.

Output sheets:
  1. 标注说明  — Instructions for annotators
  2. 人工评分表 — Scoring sheet (golden + generated, 9 blank dim columns)
  3. LLM评分参考 — LLM Judge scores (hidden from annotators during scoring)

Usage:
    # After Stage 3 completes:
    python3 scripts/human_eval_sampling.py

    # Custom sample size:
    python3 scripts/human_eval_sampling.py --sample-size 120
"""

from __future__ import annotations

import argparse
import json
import random
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[3]

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ===== Paths =====
ROOT = REPO_ROOT / 'benchmarks/psydial'
D101_PATH = Path("/root/autodl-tmp/wangqihao/datasets/PsyDial/PsyDial-D101/PsyDial-D101.json")
GENERATIONS_PATH = ROOT / "outputs" / "eval" / "d101_full" / "tppm_memory_generations.json"
SCORES_PATH = ROOT / "outputs" / "eval" / "d101_full" / "tppm_memory_judge_scores.json"
DEFAULT_OUTPUT = ROOT / "outputs" / "eval" / "d101_full" / "human_eval_scoring_qwen.xlsx"

# ===== Config =====
SAMPLE_SIZE = 100
SEED = 42

DIMS = [
    ("empathy", "共情 Emp"),
    ("active_listening", "倾听 AL"),
    ("issue_clarification", "澄清 IC"),
    ("open_ended_questioning", "开放提问 OQ"),
    ("encouraging_self_exploration", "自我探索 ESE"),
    ("cognitive_restructuring", "认知重构 CR"),
    ("guided_questioning", "引导提问 GQ"),
    ("non_judgmental_accepting_attitude", "接纳 NJAA"),
    ("overall_assessment", "综合 OA"),
]

RATING_DESC = (
    "1=差: 未展示该技能  2=较差: 技能展示不足  "
    "3=一般: 基本展示但有明显不足  4=良好: 较好展示  5=优秀: 出色展示"
)

def format_dialogue(messages):
    """Format dialogue history for spreadsheet display."""
    lines = []
    for msg in messages:
        role = msg.get("role", "")
        content = msg.get("content", "")
        if role == "user":
            lines.append(f"来访者: {content}")
        elif role == "assistant":
            lines.append(f"咨询师: {content}")
    return "\n".join(lines)

def _get_rating(entry, dim_key):
    """Extract rating from a dimension entry (handles dict or int)."""
    v = entry.get(dim_key, {})
    if isinstance(v, dict):
        return v.get("rating", 0)
    return v if isinstance(v, (int, float)) else 0

def _mean_score(entry):
    """Compute mean score across all 9 dimensions."""
    ratings = [_get_rating(entry, k) for k, _ in DIMS]
    return sum(ratings) / len(ratings) if ratings else 0

def stratified_sample(scored_cases, sample_size, seed=42):
    """Stratified sampling by mean score across all 9 dimensions."""
    random.seed(seed)

    valid = [
        c for c in scored_cases
        if c.get("generated", "").strip()
        and isinstance(c.get("overall_assessment"), dict)
    ]

    sorted_by_mean = sorted(valid, key=_mean_score)

    n = len(sorted_by_mean)
    third = n // 3

    tiers = {
        "low": sorted_by_mean[:third],
        "mid": sorted_by_mean[third : 2 * third],
        "high": sorted_by_mean[2 * third :],
    }

    per_tier = sample_size // 3
    remainder = sample_size - per_tier * 3

    sampled = []
    for i, (tier_name, tier_cases) in enumerate(tiers.items()):
        take = per_tier + (1 if i < remainder else 0)
        take = min(take, len(tier_cases))
        sampled.extend(random.sample(tier_cases, take))

    random.shuffle(sampled)
    return sampled, tiers

def build_instructions_sheet(wb):
    """Create 标注说明 sheet."""
    ws = wb.active
    ws.title = "标注说明"

    title_font = Font(bold=True, size=16)
    heading_font = Font(bold=True, size=12)
    body_font = Font(size=11)

    lines = [
        ("人工评分标注说明", title_font),
        ("", body_font),
        ("一、任务目标", heading_font),
        ("评估心理咨询场景中，咨询师回复的专业质量。每个 case 需要评估"
         "「原始回复」和「TPPM 生成回复」两个版本。", body_font),
        ("", body_font),
        ("二、评分维度（9个，1-5分 Likert 量表）", heading_font),
        ("1. 共情 Empathy (Emp) — 理解、共鸣和验证来访者情感的能力", body_font),
        ("2. 积极倾听 Active Listening (AL) — 专注倾听并准确理解来访者", body_font),
        ("3. 问题澄清 Issue Clarification (IC) — 主动寻求澄清以确保理解", body_font),
        ("4. 开放式提问 Open-ended Questioning (OQ) — 使用开放式问题鼓励探索", body_font),
        ("5. 鼓励自我探索 Encouraging Self-Exploration (ESE) — 促进来访者自我反思", body_font),
        ("6. 认知重构 Cognitive Restructuring (CR) — 帮助识别和调整不合理认知", body_font),
        ("7. 引导性提问 Guided Questioning (GQ) — 聚焦特定问题或目标进行引导", body_font),
        ("8. 非评判接纳 Non-judgmental Accepting Attitude (NJAA) — 营造安全无评判环境", body_font),
        ("9. 综合评估 Overall Assessment (OA) — 综合所有维度的整体评分", body_font),
        ("", body_font),
        (f"评分参考: {RATING_DESC}", body_font),
        ("", body_font),
        ("三、标注步骤", heading_font),
        ("1. 先阅读「对话历史」，理解咨询背景和当前进展", body_font),
        ("2. 阅读「回复文本」", body_font),
        ("3. 根据 9 个维度独立评分（1-5 整数分）", body_font),
        ("4. 如有特殊情况可在「备注」列说明", body_font),
        ("5. 每对回复（原始 / 模型生成）请独立评分，不要相互参照", body_font),
        ("", body_font),
        ("四、重要提示", heading_font),
        ("• 请基于回复文本本身的质量评分，不要因为对话轮次少就给所有维度低分", body_font),
        ("• 某些维度在特定场景下可能不适用（如认知重构在开场白中），此时给 1 分即可", body_font),
        ("• 评分应反映回复中实际展示的技能，而非回复「应该」包含什么", body_font),
    ]

    for i, (text, font) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font

    ws.column_dimensions["A"].width = 90

def build_scoring_sheet(wb, sampled, d101_index):
    """Create 人工评分表 sheet."""
    ws = wb.create_sheet("人工评分表")

    # Styles
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    golden_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    gen_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Headers
    headers = [
        "Case ID", "类型", "对话历史", "回复文本",
        "共情\n(1-5)", "倾听\n(1-5)", "澄清\n(1-5)", "开放提问\n(1-5)",
        "自我探索\n(1-5)", "认知重构\n(1-5)", "引导提问\n(1-5)",
        "接纳\n(1-5)", "综合\n(1-5)", "备注",
    ]
    col_widths = [8, 10, 55, 45, 7, 7, 7, 8, 8, 8, 8, 7, 7, 18]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    # Data rows
    row = 2
    for entry in sampled:
        idx = entry["idx"]
        d101_case = d101_index.get(idx)
        dialogue = format_dialogue(d101_case["messages"][:-1]) if d101_case else "(无对话历史)"

        for resp_type, resp_text, fill in [
            ("原始", entry.get("golden", ""), golden_fill),
            ("模型生成", entry.get("generated", ""), gen_fill),
        ]:
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

            ws.cell(row=row, column=2, value=resp_type).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

            hist_cell = ws.cell(row=row, column=3, value=dialogue)
            hist_cell.alignment = wrap
            hist_cell.border = border

            resp_cell = ws.cell(row=row, column=4, value=resp_text)
            resp_cell.alignment = wrap
            resp_cell.border = border

            for col in range(5, 15):
                c = ws.cell(row=row, column=col)
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="top")

            # Apply fill
            for col in range(1, 15):
                ws.cell(row=row, column=col).fill = fill

            ws.row_dimensions[row].height = 150
            row += 1

    # Freeze top row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = f"A1:N{row - 1}"

def build_reference_sheet(wb, sampled, d101_index):
    """Create LLM评分参考 sheet (hidden from annotators during scoring)."""
    ws = wb.create_sheet("LLM评分参考")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    # Headers
    base_headers = ["Case ID", "类型", "对话历史", "回复文本"]
    llm_headers = [f"LLM-{label.split()[1]}" for _, label in DIMS]
    human_headers = [f"人工-{label.split()[1]}" for _, label in DIMS]
    headers = base_headers + llm_headers + human_headers + ["LLM vs 人工差异备注"]

    col_widths = [8, 10, 55, 45] + [8] * len(llm_headers) + [8] * len(human_headers) + [20]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    # Data
    row = 2
    for entry in sampled:
        idx = entry["idx"]
        d101_case = d101_index.get(idx)
        dialogue = format_dialogue(d101_case["messages"][:-1]) if d101_case else ""

        for resp_type, resp_text in [("原始", entry.get("golden", "")), ("模型生成", entry.get("generated", ""))]:
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=resp_type).border = border

            hist_cell = ws.cell(row=row, column=3, value=dialogue)
            hist_cell.alignment = Alignment(wrap_text=True, vertical="top")
            hist_cell.border = border

            resp_cell = ws.cell(row=row, column=4, value=resp_text)
            resp_cell.alignment = Alignment(wrap_text=True, vertical="top")
            resp_cell.border = border

            # LLM Judge scores (only for generated response)
            if resp_type == "模型生成":
                for col_offset, (dim_key, _) in enumerate(DIMS):
                    dim = entry.get(dim_key, {})
                    rating = dim.get("rating", "") if isinstance(dim, dict) else ""
                    c = ws.cell(row=row, column=5 + col_offset, value=rating)
                    c.border = border
                    c.alignment = Alignment(horizontal="center")
            else:
                for col_offset in range(len(DIMS)):
                    ws.cell(row=row, column=5 + col_offset).border = border

            # Blank human score columns
            for col_offset in range(len(DIMS)):
                c = ws.cell(row=row, column=5 + len(DIMS) + col_offset)
                c.border = border
                c.alignment = Alignment(horizontal="center")

            # Notes column
            ws.cell(row=row, column=5 + 2 * len(DIMS)).border = border

            ws.row_dimensions[row].height = 100
            row += 1

    ws.freeze_panes = "A2"

def main():
    parser = argparse.ArgumentParser(description="Human evaluation sampling")
    parser.add_argument("--generations", type=Path, default=GENERATIONS_PATH)
    parser.add_argument("--scores", type=Path, default=SCORES_PATH)
    parser.add_argument("--d101", type=Path, default=D101_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--sample-size", type=int, default=SAMPLE_SIZE)
    parser.add_argument("--seed", type=int, default=SEED)
    args = parser.parse_args()

    # Load data
    print("Loading data...")
    with open(args.generations) as f:
        gen_data = json.load(f)

    with open(args.scores) as f:
        score_data = json.load(f)

    with open(args.d101) as f:
        d101 = json.load(f)

    d101_index = {c["idx"]: c for c in d101}
    gen_index = {r["idx"]: r for r in gen_data["results"]}

    scored_cases = score_data["scores"]
    print(f"  Generations: {len(gen_data['results'])} cases")
    print(f"  LLM Scored:  {len(scored_cases)} cases")

    if len(scored_cases) < len(gen_data["results"]) * 0.95:
        print(f"\n  ⚠ Stage 3 not fully complete ({len(scored_cases)}/{len(gen_data['results'])})")
        print(f"    Run this script again after scoring finishes.\n")

    # Stratified sample
    sampled, tiers = stratified_sample(scored_cases, args.sample_size, seed=args.seed)
    print(f"\nSampled {len(sampled)} cases (seed={args.seed}):")
    for tier_name, tier_cases in tiers.items():
        sample_in_tier = sum(1 for s in sampled if s in tier_cases)
        means = [_mean_score(s) for s in tier_cases]
        mean_str = f"{min(means):.2f}-{max(means):.2f}" if means else "N/A"
        print(f"  {tier_name:5s}: {sample_in_tier:2d} sampled / {len(tier_cases)} total (mean range: {mean_str})")

    # Build Excel
    wb = Workbook()
    build_instructions_sheet(wb)
    build_scoring_sheet(wb, sampled, d101_index)
    build_reference_sheet(wb, sampled, d101_index)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(args.output))
    print(f"\n✓ Saved: {args.output}")
    print(f"  Sheet 1: 标注说明")
    print(f"  Sheet 2: 人工评分表 ({len(sampled)} cases × 2 rows = {len(sampled)*2} rows)")
    print(f"  Sheet 3: LLM评分参考 (with LLM Judge scores for comparison)")

    # Summary stats for sampled cases
    oa_vals = [s["overall_assessment"]["rating"] for s in sampled
               if isinstance(s.get("overall_assessment"), dict)]
    print(f"\nSampled OA distribution:")
    for score in sorted(set(oa_vals)):
        count = oa_vals.count(score)
        bar = "█" * count
        print(f"  {score}: {count:2d} {bar}")
    print(f"  Mean: {sum(oa_vals)/len(oa_vals):.2f}")

if __name__ == "__main__":
    main()
