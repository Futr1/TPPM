#!/usr/bin/env python3
"""Cross-validation: deepseek-v4-pro scores the 100 sampled cases
and compares against existing LLM Judge (deepseek-chat) scores.

Uses the same scoring prompt as llm_judge_scoring.py.
Outputs per-case comparison + aggregate correlation metrics.
"""

import asyncio
import json
import os
import random
import re
import time
import argparse
from pathlib import Path
from typing import Optional

import numpy as np
from openai import AsyncOpenAI
from scipy import stats

# ===== Config =====
ROOT = Path("/root/autodl-tmp/wangqihao/Table1-data")
API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
if not API_KEY:
    raise RuntimeError(
        "DEEPSEEK_API_KEY is not set. "
        "Export it before running this script."
    )
API_BASE = os.environ.get("DEEPSEEK_API_BASE", "https://api.deepseek.com")
MODEL_NAME = os.environ.get("DEEPSEEK_JUDGE_MODEL", "deepseek-v4-pro")

GENERATIONS_PATH = ROOT / "outputs" / "eval" / "d101_full" / "tppm_memory_generations.json"
OLD_SCORES_PATH = ROOT / "outputs" / "eval" / "d101_full" / "tppm_memory_judge_scores.json"
D101_PATH = Path("/root/autodl-tmp/wangqihao/datasets/PsyDial/PsyDial-D101/PsyDial-D101.json")
HUMAN_EVAL_XLSX = ROOT / "outputs" / "eval" / "d101_full" / "human_eval_scoring_qwen.xlsx"
DEFAULT_OUTPUT = ROOT / "outputs" / "eval" / "d101_full" / "cross_validation_v4pro.json"

DIM_KEYS = [
    "empathy", "active_listening", "issue_clarification",
    "open_ended_questioning", "encouraging_self_exploration",
    "cognitive_restructuring", "guided_questioning",
    "non_judgmental_accepting_attitude", "overall_assessment",
]

DIM_ABBR = {
    "empathy": "Emp", "active_listening": "AL", "issue_clarification": "IC",
    "open_ended_questioning": "OQ", "encouraging_self_exploration": "ESE",
    "cognitive_restructuring": "CR", "guided_questioning": "GQ",
    "non_judgmental_accepting_attitude": "NJAA", "overall_assessment": "OA",
}

SYSTEM_PROMPT = (
    "You are a professional psychological counseling supervisor. "
    "You should evaluate how well the counselor applies their skills "
    "in a counseling context."
)

# Reuse exact same METRICS_DEFINITIONS from llm_judge_scoring.py
METRICS_DEFINITIONS = {
    "empathy": {
        "name": "Empathy", "abbr": "Emp",
        "definition": "Empathy refers to the counselor's ability to understand, resonate with, and validate the client's emotions and experiences. It involves not only recognizing the client's feelings but also communicating a deep sense of emotional understanding and support.",
        "criteria": "- 5: The counselor demonstrates deep empathy, consistently validating and responding to the client's emotions and experiences in a way that fosters connection.\n- 4: The counselor shows empathy, but it may occasionally lack depth or clarity in certain moments.\n- 3: The counselor shows basic empathy, but the emotional understanding feels somewhat distant or incomplete.\n- 2: The counselor struggles to demonstrate empathy, and emotional understanding feels superficial or lacking.\n- 1: The counselor does not show empathy or seems indifferent to the client's emotional experiences.",
    },
    "active_listening": {
        "name": "Active Listening", "abbr": "AL",
        "definition": "Active listening ensures a thorough understanding of the client's problems and emotions. The counselor attentively listens to both verbal and non-verbal cues, confirming the primary concerns and emotional state of the client. This helps build rapport and trust, while also making the client feel fully heard.",
        "criteria": "- 5: The counselor listens attentively without interruption, demonstrates full understanding, and accurately reflects the client's feelings and concerns.\n- 4: The counselor listens well but may occasionally miss small details or interrupt slightly.\n- 3: The counselor listens but struggles to pick up key details or misinterprets some aspects of the client's communication.\n- 2: The counselor listens partially, often missing important cues or failing to grasp the main concerns.\n- 1: The counselor does not listen actively, often interrupting or showing little engagement with the client's message.",
    },
    "issue_clarification": {
        "name": "Issue Clarification", "abbr": "IC",
        "definition": "Clarification involves seeking additional details or further explanation when the client's communication is unclear. The counselor asks specific questions to gain a better understanding of the client's situation, ensuring that all critical aspects of the problem are comprehended fully.",
        "criteria": "- 5: The counselor actively seeks clarification whenever necessary, asking precise questions that help unpack the client's issues clearly and comprehensively.\n- 4: The counselor seeks clarification in most situations, though some questions may be slightly general or unclear.\n- 3: The counselor asks some clarifying questions, but they may miss key aspects or fail to probe deeply enough.\n- 2: The counselor rarely asks for clarification, leaving gaps in understanding that could hinder progress.\n- 1: The counselor does not seek clarification, resulting in a poor understanding of the client's issues.",
    },
    "open_ended_questioning": {
        "name": "Open-ended Questioning", "abbr": "OQ",
        "definition": "Open-ended questions are designed to encourage the client to explore their thoughts, feelings, and experiences in greater depth. These questions usually start with 'how,' 'what,' or 'can you tell me more about,' and allow the client to provide expansive, reflective answers.",
        "criteria": "- 5: The counselor consistently uses open-ended questions that encourage deep exploration and self-reflection, promoting rich, meaningful dialogue.\n- 4: The counselor asks open-ended questions but may sometimes rely on closed or leading questions.\n- 3: The counselor occasionally uses open-ended questions but often defaults to yes/no questions or questions with limited scope.\n- 2: The counselor uses very few open-ended questions, limiting the client's ability to explore their own thoughts and feelings.\n- 1: The counselor avoids open-ended questions entirely, only asking yes/no or directive questions.",
    },
    "encouraging_self_exploration": {
        "name": "Encouraging Self-Exploration", "abbr": "ESE",
        "definition": "Encouraging self-exploration means asking questions and providing prompts that help the client reflect on their own emotions, thoughts, behaviors, and decision-making. This promotes greater self-awareness and empowers the client to make their own insights and choices.",
        "criteria": "- 5: The counselor frequently encourages the client to explore their own thoughts and feelings, fostering significant self-awareness and insight.\n- 4: The counselor encourages self-exploration but may not consistently prompt the client to explore deeper layers of their experiences.\n- 3: The counselor occasionally encourages self-exploration, but it may lack depth or clarity, limiting the client's reflection.\n- 2: The counselor provides limited opportunities for self-exploration, directing the conversation more than encouraging self-reflection.\n- 1: The counselor does not encourage self-exploration, instead providing solutions or interpretations without engaging the client's own thoughts.",
    },
    "cognitive_restructuring": {
        "name": "Cognitive Restructuring", "abbr": "CR",
        "definition": "Cognitive restructuring involves helping the client identify and challenge distorted or unrealistic thought patterns. The counselor assists the client in reframing negative or maladaptive thoughts, fostering more realistic and helpful cognitive patterns that promote emotional well-being.",
        "criteria": "- 5: The counselor skillfully helps the client identify distorted thoughts and gently guides them to more balanced, realistic perspectives.\n- 4: The counselor helps challenge distorted thinking, but may not consistently provide clear alternatives or insight.\n- 3: The counselor offers some cognitive restructuring, but the process feels incomplete or lacks sufficient exploration of thought patterns.\n- 2: The counselor rarely engages in cognitive restructuring, providing minimal guidance for challenging negative thoughts.\n- 1: The counselor does not address cognitive distortions or fails to help the client change unhelpful thinking patterns.",
    },
    "guided_questioning": {
        "name": "Guided Questioning", "abbr": "GQ",
        "definition": "Guided questioning refers to the use of focused questions to help the client narrow down their concerns or focus on specific goals. This approach helps the client clarify their thoughts and find solutions to specific problems, often by prompting deeper reflection on particular aspects of their experience.",
        "criteria": "- 5: The counselor uses guided questions effectively, helping the client focus on specific issues or goals in a way that enhances clarity and progress.\n- 4: The counselor uses guided questions but may not always focus them as effectively on the client's immediate needs or goals.\n- 3: The counselor uses some guiding questions, but they may be overly broad or fail to narrow in on the client's main concerns.\n- 2: The counselor rarely uses guiding questions, or their questions lack focus, making it difficult for the client to concentrate on specific issues.\n- 1: The counselor does not use guiding questions, and the session lacks focus or clarity on specific goals.",
    },
    "non_judgmental_accepting_attitude": {
        "name": "Non-judgmental and Accepting Attitude", "abbr": "NJAA",
        "definition": "A non-judgmental and accepting attitude means creating a safe and supportive environment where the client feels free to share their thoughts and feelings without fear of criticism or negative judgment. The counselor maintains a neutral, respectful approach, accepting the client's experiences and emotional expressions without imposing their own values or opinions.",
        "criteria": "- 5: The counselor consistently maintains a non-judgmental, accepting stance, allowing the client to share openly without fear of judgment.\n- 4: The counselor is generally non-judgmental, with occasional lapses in maintaining complete neutrality or acceptance.\n- 3: The counselor maintains a neutral attitude but may unintentionally come across as judgmental in some instances.\n- 2: The counselor's attitude is occasionally critical or dismissive, potentially creating discomfort for the client.\n- 1: The counselor is overtly judgmental or dismissive, making the client feel unsafe or unsupported.",
    },
    "overall_assessment": {
        "name": "Overall Assessment", "abbr": "OA",
        "definition": "This overall score combines the key evaluation principles into one holistic rating scale. The counselor's performance will be evaluated based on how effectively they apply each principle in their practice. Each principle contributes equally to the final score, providing a comprehensive assessment of the counselor's abilities in understanding and responding to the client's needs. The overall score will reflect the counselor's skill in fostering a supportive, non-judgmental, and effective therapeutic environment.",
        "evaluation_principles": "- Active Listening: Ensure a complete understanding of the client's issues and emotions, while confirming the client's main concerns and feelings.\n- Empathy: Express understanding and care for the client's emotions. Use empathetic language such as 'I understand how you feel' or 'It sounds like you're really sad.'\n- Issue Clarification: If the client's communication is unclear, the counselor can ask specific questions to ensure a full understanding of their situation.\n- Open-ended Questions: The counselor can use open-ended questions to encourage the client to provide more information and elaborate on their thoughts and feelings.\n- Encouraging Self-Exploration: The counselor can ask questions that encourage the client to explore their feelings, thoughts, and behaviors in order to promote self-reflection.\n- Cognitive Restructuring: Help the client identify and challenge unrealistic or distorted thought patterns, guiding them toward more balanced thinking.\n- Guided Questioning: The counselor can use guiding questions to help the client focus on specific issues or goals, clarifying their thoughts and moving toward resolution.\n- Non-judgmental and Accepting Attitude: Avoid making judgments about the client's experiences or emotions. Use neutral language and respect the client's perspectives and choices.",
        "criteria": "- 5: Excellent – The counselor consistently demonstrates mastery of all principles. Their approach is empathetic, insightful, and highly effective in addressing the client's needs. The counselor maintains a strong, non-judgmental rapport while promoting self-awareness and growth.\n- 4: Good – The counselor effectively applies the principles in most areas. There may be occasional lapses in one or two principles, but overall the counselor's approach is competent, and the client feels supported and understood.\n- 3: Average – The counselor demonstrates adequate skills in applying the principles. However, there are noticeable gaps or inconsistencies in their practice. Some areas may need improvement to enhance the therapeutic process.\n- 2: Below Average – The counselor struggles to effectively apply several principles. There are significant gaps in understanding or responding to the client's needs, resulting in limited therapeutic progress. The approach may sometimes feel disconnected or judgmental.\n- 1: Poor – The counselor demonstrates minimal or no proficiency in applying the principles. Their responses are ineffective, and they may create an unhelpful or even detrimental therapeutic environment. The client is likely to feel unsupported or misunderstood.",
    },
}


def format_dialogue_history(messages):
    lines = []
    for msg in messages:
        role_label = "来访者" if msg["role"] == "user" else "咨询师"
        lines.append(f"{role_label}: {msg['content']}")
    return "\n".join(lines)


def format_metric_text(key):
    m = METRICS_DEFINITIONS[key]
    text = f"{m['name']} ({m['abbr']})\n\nDefinition: {m['definition']}\n\n"
    if key == "overall_assessment" and "evaluation_principles" in m:
        text += f"Evaluation Principles:\n{m['evaluation_principles']}\n\n"
    text += f"Rating Criteria:\n{m['criteria']}"
    return text


def build_scoring_prompt(ctx, response, metric_key):
    metric_text = format_metric_text(metric_key)
    return f"""The following is a counseling context.
Dialogue history: {ctx}
Counselor's response: {response}
{metric_text}

Provide a brief reasoning for your rating based on these criteria, and then assign a numerical rating. Provide your answer in the following format.

- Reasoning: (Your explanation here)
- Rating: (Ranging from 1 to 5)"""


def parse_response(content):
    reason_match = re.search(r'- Reasoning:\s*(.+?)(?=\n- Rating:|\Z)', content, re.DOTALL)
    reasoning = reason_match.group(1).strip() if reason_match else ""
    rating_match = re.search(r'- Rating:\s*\(?Ranging from \d+ to \d+\)?\s*(\d+)', content)
    if not rating_match:
        rating_match = re.search(r'- Rating:\s*(\d+)', content)
    if rating_match:
        rating = int(rating_match.group(1))
        if 1 <= rating <= 5:
            return {"reasoning": reasoning, "rating": rating}
    return None


CONCURRENCY = 12  # max concurrent API calls across all dimensions

_clients: list = []
_client_lock = asyncio.Lock()

async def get_client():
    """Get or create an AsyncOpenAI client (one per event loop)."""
    async with _client_lock:
        if not _clients:
            _clients.append(AsyncOpenAI(api_key=API_KEY, base_url=API_BASE, timeout=120.0))
        return _clients[0]


async def call_one_dim(
    sem: asyncio.Semaphore,
    client: AsyncOpenAI,
    prompt: str,
    dim_key: str,
    max_retries: int = 3,
) -> dict:
    """Async call to score one dimension."""
    async with sem:
        for attempt in range(1, max_retries + 1):
            try:
                resp = await client.chat.completions.create(
                    model=MODEL_NAME,
                    temperature=0.0,
                    max_tokens=4096,
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": prompt},
                    ],
                )
                content = resp.choices[0].message.content or ""
                parsed = parse_response(content)
                if parsed:
                    return parsed
                if attempt < max_retries:
                    await asyncio.sleep(2)
            except Exception as e:
                if attempt >= max_retries:
                    print(f"  [{dim_key}] FAIL after {max_retries} attempts: {e}")
                    return {"reasoning": "API_FAILED", "rating": 0}
                await asyncio.sleep(min(30, 1.0 * (2 ** (attempt - 1)) + random.uniform(0, 2)))
        return {"reasoning": "API_FAILED", "rating": 0}


async def score_one_case(
    sem: asyncio.Semaphore,
    client: AsyncOpenAI,
    case_idx: int,
    ctx: str,
    response: str,
) -> dict:
    """Score all 9 dimensions concurrently for one case."""
    tasks = []
    for dim_key in DIM_KEYS:
        prompt = build_scoring_prompt(ctx, response, dim_key)
        tasks.append(call_one_dim(sem, client, prompt, dim_key))
    results = await asyncio.gather(*tasks)
    scores = {dim_key: result for dim_key, result in zip(DIM_KEYS, results)}
    scores["idx"] = case_idx
    return scores


def main():
    parser = argparse.ArgumentParser(description="Cross-validate with deepseek-v4-pro")
    parser.add_argument("--generations", type=Path, default=GENERATIONS_PATH)
    parser.add_argument("--old-scores", type=Path, default=OLD_SCORES_PATH)
    parser.add_argument("--d101", type=Path, default=D101_PATH)
    parser.add_argument("--xlsx", type=Path, default=HUMAN_EVAL_XLSX)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--delay", type=float, default=0.0)
    parser.add_argument("--concurrency", type=int, default=CONCURRENCY)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    # 1. Load sampled case IDs from Excel
    from openpyxl import load_workbook
    wb = load_workbook(args.xlsx)
    ws = wb["人工评分表"]
    sampled_ids = set()
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, 1).value
        if v is not None:
            sampled_ids.add(int(v))
    print(f"Sampled case IDs from Excel: {len(sampled_ids)} cases")

    # 2. Load data
    with open(args.generations) as f:
        gen_data = json.load(f)
    gen_index = {r["idx"]: r for r in gen_data["results"]}

    with open(args.old_scores) as f:
        old_scores_data = json.load(f)
    old_scores_index = {s["idx"]: s for s in old_scores_data["scores"]}

    with open(args.d101) as f:
        d101 = json.load(f)
    d101_index = {c["idx"]: c for c in d101}

    # 3. Filter to sampled cases with valid generations
    cases_to_score = []
    for idx in sorted(sampled_ids):
        gen = gen_index.get(idx)
        if gen and gen.get("generated", "").strip():
            cases_to_score.append(gen)
    print(f"Cases to score: {len(cases_to_score)}")

    # 4. Load checkpoint
    existing = {}
    if args.output.exists():
        with open(args.output) as f:
            existing = {s["idx"]: s for s in json.load(f).get("scores", [])}
        print(f"Existing checkpoint: {len(existing)} cases")

    if args.dry_run:
        print("\n=== DRY RUN ===")
        c = cases_to_score[0]
        d101_case = d101_index.get(c["idx"])
        msgs = d101_case["messages"][:-1] if d101_case else []
        ctx = format_dialogue_history(msgs) if msgs else "(no history)"
        prompt = build_scoring_prompt(ctx, c["generated"], "empathy")
        print(f"Case idx={c['idx']}")
        print(prompt[:500])
        return

    # 5. Score each case with deepseek-v4-pro (async, concurrent dims)
    scores_list = list(existing.values())
    total = len(cases_to_score)
    start_time = time.time()

    # Pre-build prompts for all cases (to avoid repeated work in the loop)
    for gen in cases_to_score:
        idx = gen["idx"]
        if idx in existing and all(k in existing[idx] for k in DIM_KEYS):
            continue
        d101_case = d101_index.get(idx)
        msgs = d101_case["messages"][:-1] if d101_case else []
        gen["_ctx"] = format_dialogue_history(msgs) if msgs else "(no history)"

    async def run_async():
        nonlocal scores_list
        client = await get_client()
        sem = asyncio.Semaphore(args.concurrency)
        scored = len(scores_list)

        pending = [g for g in cases_to_score
                   if g["idx"] not in existing or not all(k in existing[g["idx"]] for k in DIM_KEYS)]

        for gen in pending:
            idx = gen["idx"]
            ctx = gen.get("_ctx", "(no history)")
            response = gen["generated"]

            entry = await score_one_case(sem, client, idx, ctx, response)
            entry["golden"] = gen.get("golden", "")
            entry["generated"] = response
            scores_list.append(entry)

            elapsed = time.time() - start_time
            scored += 1
            rate = scored / (elapsed / 60) if elapsed > 0 else 0
            ratings = " ".join(f"{DIM_ABBR[k]}={entry[k]['rating']}" for k in DIM_KEYS)
            print(f"[{scored}/{len(pending)}] idx={idx} | {ratings} | {rate:.1f}/min")

            # Checkpoint every 5 cases
            if scored % 5 == 0:
                args.output.parent.mkdir(parents=True, exist_ok=True)
                with open(args.output, "w") as f:
                    json.dump({
                        "metadata": {"scorer": MODEL_NAME, "target": "cross_validation",
                                     "sampled_from": "human_eval_scoring_qwen.xlsx"},
                        "scores": sorted(scores_list, key=lambda x: x["idx"]),
                    }, f, ensure_ascii=False, indent=2)

    asyncio.run(run_async())

    # Final save
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w") as f:
        json.dump({
            "metadata": {"scorer": MODEL_NAME, "target": "cross_validation",
                         "sampled_from": "human_eval_scoring_qwen.xlsx"},
            "scores": sorted(scores_list, key=lambda x: x["idx"]),
        }, f, ensure_ascii=False, indent=2)
    print(f"\nSaved: {args.output}")

    # 6. Cross-validation: compare v4-pro vs old LLM Judge
    print("\n" + "=" * 70)
    print("CROSS-VALIDATION: deepseek-v4-pro vs old LLM Judge (deepseek-chat)")
    print("=" * 70)

    v4pro_index = {s["idx"]: s for s in scores_list}
    matched = []
    for idx in sorted(sampled_ids):
        old = old_scores_index.get(idx)
        new = v4pro_index.get(idx)
        if old and new:
            matched.append((idx, old, new))

    print(f"Matched cases: {len(matched)}")

    # Per-dimension correlation
    print(f"\n{'Dimension':<20} {'Pearson r':>10} {'Spearman ρ':>10} {'MAE':>6} {'Exact%':>7} {'Δ≤1':>6}")
    print("-" * 70)

    all_old, all_new = [], []
    for dim_key in DIM_KEYS:
        old_vals, new_vals = [], []
        for idx, old, new in matched:
            ov = old.get(dim_key, {})
            nv = new.get(dim_key, {})
            o_rating = ov.get("rating", 0) if isinstance(ov, dict) else ov
            n_rating = nv.get("rating", 0) if isinstance(nv, dict) else nv
            if o_rating > 0 and n_rating > 0:
                old_vals.append(o_rating)
                new_vals.append(n_rating)

        if len(old_vals) < 3:
            print(f"{DIM_ABBR[dim_key]:<20} {'N/A':>10} {'N/A':>10} {'N/A':>6} {'N/A':>7} {'N/A':>6}")
            continue

        r, _ = stats.pearsonr(old_vals, new_vals)
        rho, _ = stats.spearmanr(old_vals, new_vals)
        mae = np.mean(np.abs(np.array(old_vals) - np.array(new_vals)))
        exact_pct = sum(1 for a, b in zip(old_vals, new_vals) if a == b) / len(old_vals) * 100
        within1_pct = sum(1 for a, b in zip(old_vals, new_vals) if abs(a - b) <= 1) / len(old_vals) * 100

        print(f"{DIM_ABBR[dim_key]:<20} {r:>10.3f} {rho:>10.3f} {mae:>6.2f} {exact_pct:>6.1f}% {within1_pct:>5.1f}%")

        all_old.extend(old_vals)
        all_new.extend(new_vals)

    # Overall
    if all_old:
        r, _ = stats.pearsonr(all_old, all_new)
        rho, _ = stats.spearmanr(all_old, all_new)
        mae = np.mean(np.abs(np.array(all_old) - np.array(all_new)))
        exact_pct = sum(1 for a, b in zip(all_old, all_new) if a == b) / len(all_old) * 100
        within1_pct = sum(1 for a, b in zip(all_old, all_new) if abs(a - b) <= 1) / len(all_old) * 100
        print("-" * 70)
        print(f"{'ALL':<20} {r:>10.3f} {rho:>10.3f} {mae:>6.2f} {exact_pct:>6.1f}% {within1_pct:>5.1f}%")


if __name__ == "__main__":
    main()
